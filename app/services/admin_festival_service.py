from datetime import date, datetime
from io import BytesIO
import json
from pathlib import Path
import re
import unicodedata

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import xlrd

from app.core.firebase import db
from app.core.utils import serialize_date, utc_now_iso
from app.schemas.festival_schema import (
    FestivalResponse,
    FestivalStatus,
    FestivalUpdateRequest,
)


FESTIVALS_COLLECTION = "festivals"
FESTIVAL_CLEANUP_LOGS_COLLECTION = "festival_cleanup_logs"
AUDIT_FILE = Path(__file__).resolve().parents[2] / "festival-audit.json"
DATE_FIELDS = {"opening_date", "deadline", "event_date"}
COMPLETENESS_FIELDS = (
    "name",
    "country",
    "website",
    "submission_url",
    "platform",
    "opening_date",
    "deadline",
    "event_date",
    "fee",
    "edition_year",
    "contact",
    "notes",
)
MERGEABLE_FIELDS = COMPLETENESS_FIELDS[1:] + ("form_fields",)
EDITABLE_FIELDS = {
    "name",
    "country",
    "website",
    "submission_url",
    "platform",
    "opening_date",
    "deadline",
    "event_date",
    "fee",
    "status",
    "edition_year",
    "contact",
    "notes",
}


def _normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_key(value) -> str:
    text = unicodedata.normalize("NFKD", _normalize_text(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


HEADER_ALIASES = {
    "name": {
        "festival",
        "festival en verde convocatorias abiertas",
        "nombre",
        "nombre festival",
        "festivales",
        "name",
    },
    "country": {"pais", "country"},
    "status": {"convocatoria", "estado", "estado convocatoria", "status"},
    "website": {"web", "website", "url", "link"},
    "submission_url": {
        "url postulacion",
        "submission url",
        "submission_url",
    },
    "platform": {"plataforma", "platform"},
    "opening_date": {
        "fechas inicio postulaciones",
        "fecha apertura convocatoria",
        "apertura",
        "opening date",
        "opening_date",
    },
    "deadline": {"deadline", "fecha cierre", "cierre"},
    "event_date": {"fecha evento", "event date", "event_date"},
    "fee": {"tasa", "fee", "early bird", "estandar", "late"},
    "contact": {"contacto", "contact"},
    "notes": {
        "comments",
        "comentarios",
        "waiver",
        "waiver solicitado",
        "notes",
        "notas",
    },
    "edition_year": {
        "edition year",
        "edition_year",
        "ano",
        "ano edicion",
        "edicion",
    },
}

ALIAS_TO_FIELD = {
    alias: field
    for field, aliases in HEADER_ALIASES.items()
    for alias in aliases
}

STATUS_ALIASES = {
    "open": FestivalStatus.OPEN.value,
    "abierta": FestivalStatus.OPEN.value,
    "abierto": FestivalStatus.OPEN.value,
    "upcoming": FestivalStatus.UPCOMING.value,
    "proximamente": FestivalStatus.UPCOMING.value,
    "closed": FestivalStatus.CLOSED.value,
    "cerrada": FestivalStatus.CLOSED.value,
    "cerrado": FestivalStatus.CLOSED.value,
    "archived": FestivalStatus.ARCHIVED.value,
    "archivada": FestivalStatus.ARCHIVED.value,
    "archivado": FestivalStatus.ARCHIVED.value,
    "unknown": FestivalStatus.UNKNOWN.value,
    "desconocido": FestivalStatus.UNKNOWN.value,
}


class _LegacySheet:
    def __init__(self, sheet, datemode: int):
        self._sheet = sheet
        self._datemode = datemode
        self.title = sheet.name
        self.max_row = sheet.nrows

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = True,
    ):
        upper_bound = min(max_row or self._sheet.nrows, self._sheet.nrows)
        for row_index in range(max(min_row - 1, 0), upper_bound):
            values = []
            for cell in self._sheet.row(row_index):
                if cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(
                        xlrd.xldate_as_datetime(cell.value, self._datemode)
                    )
                elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                    values.append(None)
                else:
                    values.append(cell.value)
            yield tuple(values)


class _LegacyWorkbook:
    def __init__(self, workbook):
        self._workbook = workbook
        self.epoch = None
        self.worksheets = [
            _LegacySheet(sheet, workbook.datemode)
            for sheet in workbook.sheets()
        ]

    def close(self):
        self._workbook.release_resources()


def _load_excel_workbook(file_bytes: bytes):
    try:
        return load_workbook(
            filename=BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception:
        try:
            return _LegacyWorkbook(xlrd.open_workbook(file_contents=file_bytes))
        except Exception as xlrd_error:
            raise ValueError(
                "No se pudo leer el archivo. Debe ser un libro Excel .xlsx o .xls valido."
            ) from xlrd_error


def _normalize_date(value, epoch=None) -> str:
    if value is None or _normalize_text(value) == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch=epoch)
            return (
                converted.date().isoformat()
                if isinstance(converted, datetime)
                else converted.isoformat()
            )
        except (TypeError, ValueError, OverflowError) as exc:
            raise ValueError(f"fecha Excel invalida: {value}") from exc

    raw_value = _normalize_text(value)
    for date_format in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw_value, date_format).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"formato de fecha no soportado: {raw_value}")


def _parse_date(value) -> date | None:
    normalized = _normalize_date(value)
    return date.fromisoformat(normalized) if normalized else None


def _normalize_status(value) -> str:
    normalized = _normalize_key(value)
    if not normalized:
        return FestivalStatus.UNKNOWN.value
    status = STATUS_ALIASES.get(normalized, normalized.upper())
    try:
        return FestivalStatus(status).value
    except ValueError:
        return FestivalStatus.UNKNOWN.value


def _calculate_status(
    opening_date,
    deadline,
    current_status=None,
    today: date | None = None,
) -> str:
    if _normalize_status(current_status) == FestivalStatus.ARCHIVED.value:
        return FestivalStatus.ARCHIVED.value

    current_date = today or date.today()
    opening = _parse_date(opening_date)
    closing = _parse_date(deadline)

    if closing and closing < current_date:
        return FestivalStatus.CLOSED.value
    if opening and opening > current_date:
        return FestivalStatus.UPCOMING.value
    if opening and closing and opening <= current_date <= closing:
        return FestivalStatus.OPEN.value
    if closing and closing >= current_date:
        return FestivalStatus.OPEN.value
    return FestivalStatus.UNKNOWN.value


def _serialize_festival(festival_id: str, data: dict) -> FestivalResponse:
    return FestivalResponse(
        id=data.get("id") or festival_id,
        name=_normalize_text(data.get("name")),
        country=_normalize_text(data.get("country")),
        website=_normalize_text(data.get("website")),
        submission_url=_normalize_text(data.get("submission_url")),
        platform=_normalize_text(data.get("platform")),
        opening_date=serialize_date(data.get("opening_date")) or "",
        deadline=serialize_date(data.get("deadline")) or "",
        event_date=serialize_date(data.get("event_date")) or "",
        fee=_normalize_text(data.get("fee")),
        status=_normalize_status(data.get("status")),
        form_fields=data.get("form_fields") or [],
        edition_year=_normalize_text(data.get("edition_year")),
        contact=_normalize_text(data.get("contact")),
        notes=_normalize_text(data.get("notes")),
        source=_normalize_text(data.get("source")) or "excel",
        last_checked_at=serialize_date(data.get("last_checked_at")) or "",
        created_at=serialize_date(data.get("created_at")) or "",
        updated_at=serialize_date(data.get("updated_at")) or "",
    )


def _dedupe_key(name, country, edition_year) -> tuple[str, str, str]:
    return (
        _normalize_key(name),
        _normalize_key(country),
        _normalize_key(edition_year),
    )


def _get_safe_max_row(sheet, fallback: int = 30) -> int:
    max_row = getattr(sheet, "max_row", None)
    if isinstance(max_row, int) and max_row > 0:
        return min(max_row, fallback)
    return fallback


def _get_sheet_row_count(sheet) -> int:
    for attribute in ("max_row", "nrows"):
        row_count = getattr(sheet, attribute, None)
        if isinstance(row_count, int) and row_count > 0:
            return row_count
    return 0


def _iter_sheet_rows(
    sheet,
    max_rows: int | None = 30,
    min_row: int = 1,
):
    iter_rows = getattr(sheet, "iter_rows", None)
    if callable(iter_rows):
        kwargs = {
            "min_row": min_row,
            "values_only": True,
        }
        if max_rows is not None:
            kwargs["max_row"] = _get_safe_max_row(sheet, max_rows)
        yield from iter_rows(**kwargs)
        return

    row_values = getattr(sheet, "row_values", None)
    nrows = getattr(sheet, "nrows", None)
    if isinstance(nrows, int) and nrows > 0 and callable(row_values):
        start_index = max(min_row - 1, 0)
        upper_bound = nrows if max_rows is None else min(nrows, max_rows)
        for row_index in range(start_index, upper_bound):
            yield tuple(row_values(row_index))


def _find_header_row(sheet) -> tuple[int, dict[str, list[int]], dict[int, str]] | None:
    for row_number, row in enumerate(
        _iter_sheet_rows(sheet, 30),
        start=1,
    ):
        fields: dict[str, list[int]] = {}
        labels: dict[int, str] = {}
        for index, value in enumerate(row):
            normalized = _normalize_key(value)
            field = ALIAS_TO_FIELD.get(normalized)
            if field:
                fields.setdefault(field, []).append(index)
                labels[index] = _normalize_text(value)
        if "name" in fields:
            return row_number, fields, labels
    return None


def _first_value(row: tuple, indices: list[int]) -> object:
    for index in indices:
        if index < len(row) and _normalize_text(row[index]):
            return row[index]
    return ""


def _combined_value(
    row: tuple,
    indices: list[int],
    labels: dict[int, str],
) -> str:
    values = [
        (labels.get(index, ""), _normalize_text(row[index]))
        for index in indices
        if index < len(row) and _normalize_text(row[index])
    ]
    if len(values) == 1:
        return values[0][1]
    return " | ".join(f"{label}: {value}" for label, value in values)


def _row_to_festival(
    row: tuple,
    fields: dict[str, list[int]],
    labels: dict[int, str],
    epoch,
) -> dict:
    data = {}
    for field, indices in fields.items():
        if field in {"fee", "notes"}:
            data[field] = _combined_value(row, indices, labels)
        else:
            data[field] = _first_value(row, indices)

    for field in DATE_FIELDS:
        data[field] = _normalize_date(data.get(field), epoch=epoch)

    for field in EDITABLE_FIELDS - DATE_FIELDS - {"status"}:
        data[field] = _normalize_text(data.get(field))

    if not data.get("edition_year"):
        reference_date = (
            data.get("event_date")
            or data.get("deadline")
            or data.get("opening_date")
        )
        data["edition_year"] = reference_date[:4] if reference_date else ""

    incoming_status = _normalize_status(data.get("status"))
    data["status"] = _calculate_status(
        data.get("opening_date"),
        data.get("deadline"),
        incoming_status,
    )
    return data


def import_festivals_from_excel(file_bytes: bytes) -> dict:
    if not file_bytes:
        raise ValueError("El archivo Excel esta vacio")

    workbook = _load_excel_workbook(file_bytes)

    existing_by_key = {}
    for snapshot in db.collection(FESTIVALS_COLLECTION).stream():
        data = snapshot.to_dict() or {}
        existing_by_key[
            _dedupe_key(
                data.get("name"),
                data.get("country"),
                data.get("edition_year"),
            )
        ] = (snapshot.reference, data)

    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    try:
        for sheet in workbook.worksheets:
            header = _find_header_row(sheet)
            if header is None:
                result["skipped"] += _get_sheet_row_count(sheet)
                result["errors"].append(
                    f"Hoja '{sheet.title}': no se encontro una columna de nombre"
                )
                continue

            header_row, fields, labels = header
            for row_number, row in enumerate(
                _iter_sheet_rows(
                    sheet,
                    max_rows=None,
                    min_row=header_row + 1,
                ),
                start=header_row + 1,
            ):
                try:
                    data = _row_to_festival(row, fields, labels, workbook.epoch)
                    if not data.get("name"):
                        result["skipped"] += 1
                        continue

                    key = _dedupe_key(
                        data["name"],
                        data.get("country"),
                        data.get("edition_year"),
                    )
                    timestamp = utc_now_iso()
                    existing = existing_by_key.get(key)
                    if existing:
                        document_ref, existing_data = existing
                        festival_data = {
                            **existing_data,
                            **data,
                            "id": existing_data.get("id") or document_ref.id,
                            "form_fields": existing_data.get("form_fields") or [],
                            "source": "excel",
                            "last_checked_at": existing_data.get("last_checked_at") or "",
                            "created_at": existing_data.get("created_at") or timestamp,
                            "updated_at": timestamp,
                        }
                        document_ref.set(festival_data)
                        existing_by_key[key] = (document_ref, festival_data)
                        result["updated"] += 1
                    else:
                        document_ref = db.collection(FESTIVALS_COLLECTION).document()
                        festival_data = {
                            **data,
                            "id": document_ref.id,
                            "form_fields": [],
                            "source": "excel",
                            "last_checked_at": "",
                            "created_at": timestamp,
                            "updated_at": timestamp,
                        }
                        document_ref.set(festival_data)
                        existing_by_key[key] = (document_ref, festival_data)
                        result["created"] += 1
                except Exception as exc:
                    result["skipped"] += 1
                    result["errors"].append(
                        f"Hoja '{sheet.title}', fila {row_number}: {exc}"
                    )
    finally:
        workbook.close()

    return result


def list_admin_festivals(
    status: str | None = None,
    country: str | None = None,
    search: str | None = None,
    platform: str | None = None,
    limit: int = 500,
) -> list[FestivalResponse]:
    requested_status = _normalize_status(status) if status else None
    country_filter = _normalize_key(country)
    platform_filter = _normalize_key(platform)
    search_filter = _normalize_key(search)
    items = []

    for snapshot in db.collection(FESTIVALS_COLLECTION).stream():
        data = snapshot.to_dict() or {}
        if requested_status and _normalize_status(data.get("status")) != requested_status:
            continue
        if country_filter and country_filter not in _normalize_key(data.get("country")):
            continue
        if platform_filter and platform_filter not in _normalize_key(data.get("platform")):
            continue
        if search_filter:
            searchable = " ".join(
                _normalize_key(data.get(field))
                for field in ("name", "country", "platform", "website")
            )
            if search_filter not in searchable:
                continue
        items.append(_serialize_festival(snapshot.id, data))

    items.sort(key=lambda item: (not bool(item.deadline), item.deadline, item.name.lower()))
    return items[:limit]


def update_admin_festival(
    festival_id: str,
    payload: FestivalUpdateRequest,
) -> FestivalResponse:
    document_ref = db.collection(FESTIVALS_COLLECTION).document(festival_id)
    snapshot = document_ref.get()
    if not snapshot.exists:
        raise HTTPException(status_code=404, detail="Festival no encontrado")

    existing_data = snapshot.to_dict() or {}
    updates = payload.model_dump(exclude_unset=True)
    for field, value in list(updates.items()):
        if field in DATE_FIELDS:
            updates[field] = _normalize_date(value)
        elif field == "status":
            updates[field] = (
                value.value if isinstance(value, FestivalStatus) else _normalize_status(value)
            )
        else:
            updates[field] = _normalize_text(value)

    updated_data = {
        **existing_data,
        **updates,
        "id": existing_data.get("id") or festival_id,
        "form_fields": existing_data.get("form_fields") or [],
        "source": existing_data.get("source") or "excel",
        "created_at": existing_data.get("created_at") or utc_now_iso(),
        "updated_at": utc_now_iso(),
    }
    if "status" not in updates:
        updated_data["status"] = _calculate_status(
            updated_data.get("opening_date"),
            updated_data.get("deadline"),
            existing_data.get("status"),
        )

    document_ref.set(updated_data)
    return _serialize_festival(festival_id, updated_data)


def refresh_festival_statuses() -> dict:
    counts = {status.value: 0 for status in FestivalStatus}
    updated = 0

    for snapshot in db.collection(FESTIVALS_COLLECTION).stream():
        data = snapshot.to_dict() or {}
        old_status = _normalize_status(data.get("status"))
        new_status = _calculate_status(
            data.get("opening_date"),
            data.get("deadline"),
            old_status,
        )
        counts[new_status] += 1
        if new_status != old_status:
            snapshot.reference.set(
                {"status": new_status, "updated_at": utc_now_iso()},
                merge=True,
            )
            updated += 1

    return {"updated": updated, "counts": counts}


def _load_festival_audit() -> dict:
    try:
        with AUDIT_FILE.open(encoding="utf-8") as audit_file:
            return json.load(audit_file)
    except FileNotFoundError as exc:
        raise HTTPException(
            status_code=503,
            detail="No se encontro festival-audit.json",
        ) from exc
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=503,
            detail="festival-audit.json no contiene JSON valido",
        ) from exc


def _audit_firestore() -> dict:
    firestore_audit = _load_festival_audit().get("firestore")
    if not isinstance(firestore_audit, dict):
        raise HTTPException(
            status_code=503,
            detail="festival-audit.json no contiene la seccion firestore",
        )
    return firestore_audit


def _is_present(value) -> bool:
    if value is None or value is False:
        return False
    if isinstance(value, str):
        return bool(value.strip()) and value.strip() not in {"-", "0"}
    if isinstance(value, (list, dict, tuple, set)):
        return bool(value)
    return True


def _completeness_score(data: dict) -> int:
    return sum(1 for field in COMPLETENESS_FIELDS if _is_present(data.get(field)))


def _audit_document_ids() -> set[str]:
    audit = _audit_firestore()
    return {
        item["id"]
        for group in audit.get("semantic_duplicate_groups", [])
        for item in group
        if item.get("id")
    } | {
        item["id"]
        for item in audit.get("invalid_documents", [])
        if item.get("id")
    }


def _get_audited_documents() -> dict[str, dict]:
    document_ids = sorted(_audit_document_ids())
    references = [
        db.collection(FESTIVALS_COLLECTION).document(document_id)
        for document_id in document_ids
    ]
    documents = {}
    for snapshot in db.get_all(references):
        if snapshot.exists:
            documents[snapshot.id] = snapshot.to_dict() or {}
    return documents


def _public_cleanup_document(
    document_id: str,
    data: dict,
    *,
    recommended_keep: bool = False,
    recommended_delete: bool = False,
) -> dict:
    return {
        "id": document_id,
        "name": _normalize_text(data.get("name")),
        "country": _normalize_text(data.get("country")),
        "website": _normalize_text(data.get("website")),
        "deadline": serialize_date(data.get("deadline")) or "",
        "status": _normalize_status(data.get("status")),
        "edition_year": _normalize_text(data.get("edition_year")),
        "completeness_score": _completeness_score(data),
        "recommended_keep": recommended_keep,
        "recommended_delete": recommended_delete,
    }


def _build_duplicate_plan(documents: dict[str, dict] | None = None) -> dict:
    audit = _audit_firestore()
    live_documents = documents if documents is not None else _get_audited_documents()
    groups = []
    keep_documents = []
    archive_documents = []
    merge_operations = []
    blocked_groups = []

    for audit_group in audit.get("semantic_duplicate_groups", []):
        existing = [
            (item["id"], live_documents[item["id"]])
            for item in audit_group
            if item.get("id") in live_documents
        ]
        active = [
            (document_id, data)
            for document_id, data in existing
            if not (
                _normalize_status(data.get("status")) == FestivalStatus.ARCHIVED.value
                and data.get("archived_reason") == "DUPLICATE"
            )
        ]
        ranked = sorted(
            active,
            key=lambda item: (-_completeness_score(item[1]), item[0]),
        )
        winner_id = ranked[0][0] if ranked else None
        top_score = _completeness_score(ranked[0][1]) if ranked else 0
        tied_ids = [
            document_id
            for document_id, data in ranked
            if _completeness_score(data) == top_score
        ]
        edition_years = {
            _normalize_text(data.get("edition_year"))
            for _, data in active
            if _normalize_text(data.get("edition_year"))
        }

        blocked_reason = None
        if len(active) < 2:
            blocked_reason = "ALREADY_CLEAN_OR_DOCUMENT_MISSING"
        elif len(edition_years) > 1:
            blocked_reason = "DISTINCT_EDITION_YEARS"
        elif len(tied_ids) > 1:
            blocked_reason = "AMBIGUOUS_COMPLETENESS_TIE"

        safe = blocked_reason is None
        canonical_name = (
            _normalize_text(ranked[0][1].get("name"))
            if ranked
            else _normalize_text(audit_group[0].get("name")) if audit_group else ""
        )
        public_documents = [
            _public_cleanup_document(
                document_id,
                data,
                recommended_keep=safe and document_id == winner_id,
                recommended_delete=safe and document_id != winner_id,
            )
            for document_id, data in sorted(existing)
        ]
        group_result = {
            "canonical_name": canonical_name,
            "documents": public_documents,
            "cleanup_safe": safe,
            "blocked_reason": blocked_reason,
        }
        groups.append(group_result)

        if not safe:
            if blocked_reason != "ALREADY_CLEAN_OR_DOCUMENT_MISSING":
                blocked_groups.append(group_result)
            continue

        winner_data = live_documents[winner_id]
        losers = [
            (document_id, data)
            for document_id, data in active
            if document_id != winner_id
        ]
        merged_fields = {}
        merge_sources = {}
        for field in MERGEABLE_FIELDS:
            if _is_present(winner_data.get(field)):
                continue
            for loser_id, loser_data in losers:
                if _is_present(loser_data.get(field)):
                    merged_fields[field] = loser_data[field]
                    merge_sources[field] = loser_id
                    break

        keep_documents.append(
            _public_cleanup_document(
                winner_id,
                winner_data,
                recommended_keep=True,
            )
        )
        for loser_id, loser_data in losers:
            archive_documents.append(
                _public_cleanup_document(
                    loser_id,
                    loser_data,
                    recommended_delete=True,
                )
            )
        merge_operations.append(
            {
                "target_id": winner_id,
                "source_ids": [document_id for document_id, _ in losers],
                "fields": merged_fields,
                "field_sources": merge_sources,
            }
        )

    return {
        "groups": groups,
        "keep_documents": keep_documents,
        "archive_documents": archive_documents,
        "merge_operations": merge_operations,
        "blocked_groups": blocked_groups,
    }


def get_festival_audit_summary() -> dict:
    audit = _audit_firestore()
    status_counts = {
        status.value: int(audit.get("status_raw_all_documents", {}).get(status.value, 0))
        for status in FestivalStatus
    }
    return {
        "total_documents": int(audit.get("total_documents_raw", 0)),
        "valid_unique_festivals": int(audit.get("unique_semantic_names", 0)),
        "duplicate_documents": int(
            audit.get("duplicates_semantic_name", {}).get("duplicate_rows", 0)
        ),
        "invalid_auxiliary_documents": len(audit.get("invalid_documents", [])),
        "missing_from_firestore": int(audit.get("missing_unique_count", 0)),
        "incomplete_documents": (
            int(audit.get("incomplete_records", 0))
            + len(audit.get("invalid_documents", []))
        ),
        "status_counts": status_counts,
    }


def list_festival_duplicates() -> list[dict]:
    return _build_duplicate_plan()["groups"]


def preview_festival_cleanup() -> dict:
    live_documents = _get_audited_documents()
    plan = _build_duplicate_plan(live_documents)
    invalid_documents = [
        _public_cleanup_document(item["id"], live_documents[item["id"]])
        for item in _audit_firestore().get("invalid_documents", [])
        if item.get("id") in live_documents
    ]
    total_documents = int(_audit_firestore().get("total_documents_raw", 0))
    archive_count = len(plan["archive_documents"])
    return {
        "documents_to_keep": plan["keep_documents"],
        "documents_to_archive": plan["archive_documents"],
        "invalid_auxiliary_documents": invalid_documents,
        "documents_to_merge": plan["merge_operations"],
        "blocked_groups": plan["blocked_groups"],
        "estimated_count_changes": {
            "physical_documents_before": total_documents,
            "physical_documents_after": total_documents,
            "active_documents_before": total_documents,
            "active_documents_after_duplicate_cleanup": total_documents - archive_count,
            "duplicates_archived": archive_count,
            "invalid_documents_pending": len(invalid_documents),
        },
    }


def _require_cleanup_confirmation(confirm: bool) -> None:
    if confirm is not True:
        raise HTTPException(
            status_code=400,
            detail="La limpieza requiere confirm=true",
        )


def cleanup_duplicate_festivals(confirm: bool, actor_uid: str) -> dict:
    _require_cleanup_confirmation(confirm)
    documents = _get_audited_documents()
    plan = _build_duplicate_plan(documents)
    timestamp = utc_now_iso()
    batch = db.batch()

    for operation in plan["merge_operations"]:
        target_id = operation["target_id"]
        updates = {
            **operation["fields"],
            "updated_at": timestamp,
            "merged_duplicate_ids": operation["source_ids"],
        }
        batch.set(
            db.collection(FESTIVALS_COLLECTION).document(target_id),
            updates,
            merge=True,
        )
        for source_id in operation["source_ids"]:
            batch.set(
                db.collection(FESTIVALS_COLLECTION).document(source_id),
                {
                    "status": FestivalStatus.ARCHIVED.value,
                    "archived_reason": "DUPLICATE",
                    "merged_into": target_id,
                    "archived_at": timestamp,
                    "updated_at": timestamp,
                },
                merge=True,
            )

    log_ref = db.collection(FESTIVAL_CLEANUP_LOGS_COLLECTION).document()
    log_data = {
        "id": log_ref.id,
        "action": "CLEANUP_DUPLICATES",
        "actor_uid": actor_uid,
        "created_at": timestamp,
        "archived_count": len(plan["archive_documents"]),
        "merged_group_count": len(plan["merge_operations"]),
        "blocked_group_count": len(plan["blocked_groups"]),
        "archived_document_ids": [
            item["id"] for item in plan["archive_documents"]
        ],
        "merges": plan["merge_operations"],
    }
    batch.set(log_ref, log_data)
    batch.commit()
    refreshed = refresh_festival_statuses()

    return {
        "cleanup_log_id": log_ref.id,
        "archived_count": len(plan["archive_documents"]),
        "merged_group_count": len(plan["merge_operations"]),
        "blocked_groups": plan["blocked_groups"],
        "status_refresh": refreshed,
    }


def cleanup_invalid_festivals(confirm: bool, actor_uid: str) -> dict:
    _require_cleanup_confirmation(confirm)
    audit = _audit_firestore()
    documents = _get_audited_documents()
    invalid_ids = [
        item["id"]
        for item in audit.get("invalid_documents", [])
        if item.get("id") in documents
        and not (
            _normalize_status(documents[item["id"]].get("status"))
            == FestivalStatus.ARCHIVED.value
            and documents[item["id"]].get("archived_reason")
            == "INVALID_AUXILIARY_ROW"
        )
    ]
    timestamp = utc_now_iso()
    batch = db.batch()
    for document_id in invalid_ids:
        batch.set(
            db.collection(FESTIVALS_COLLECTION).document(document_id),
            {
                "status": FestivalStatus.ARCHIVED.value,
                "archived_reason": "INVALID_AUXILIARY_ROW",
                "archived_at": timestamp,
                "updated_at": timestamp,
            },
            merge=True,
        )

    log_ref = db.collection(FESTIVAL_CLEANUP_LOGS_COLLECTION).document()
    batch.set(
        log_ref,
        {
            "id": log_ref.id,
            "action": "CLEANUP_INVALID_AUXILIARY",
            "actor_uid": actor_uid,
            "created_at": timestamp,
            "archived_count": len(invalid_ids),
            "archived_document_ids": invalid_ids,
        },
    )
    batch.commit()
    refreshed = refresh_festival_statuses()
    return {
        "cleanup_log_id": log_ref.id,
        "archived_count": len(invalid_ids),
        "archived_document_ids": invalid_ids,
        "status_refresh": refreshed,
    }
