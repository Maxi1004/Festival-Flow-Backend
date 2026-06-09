from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
import json
from pathlib import Path
import re
import sys
import unicodedata
from urllib.parse import urlparse

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


DATE_FORMATS = (
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%Y-%m-%d",
)
MONTH_NAMES = {
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "setiembre",
    "octubre",
    "noviembre",
    "diciembre",
    "january",
    "february",
    "march",
    "april",
    "may",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
}
GENERIC_NAME_WORDS = {
    "festival",
    "fest",
    "film",
    "films",
    "cine",
    "cinema",
    "international",
    "internacional",
    "de",
    "del",
    "la",
    "el",
    "of",
    "the",
}
HEADER_NAMES = {
    "festival",
    "festivales",
    "nombre festival",
    "festival en verde convocatorias abiertas",
}
STATUS_VALUES = {"OPEN", "UPCOMING", "CLOSED", "ARCHIVED", "UNKNOWN"}
INVALID_NAME_KEYS = {
    "2025",
    "cyprus",
    "ciclo de cine y salud mental no veo que tengan web ni nada te paso lo que he visto",
}
ALIAS_GROUPS = (
    (
        "Festival Internacional de Cine de Mar del Plata",
        "Festival Internacional de Cine de Mar de Plata",
    ),
    ("Festival de Málaga", "Festival de Cine de Málaga"),
    (
        "Festival Internacional de Cine de Valdivia",
        "Festival de Cine de Valdivia",
    ),
    ("Doclisboa - International Film Festival", "Doclisboa Film Festival"),
    ("International Istanbul Film Festival", "Istanbul Film Festival"),
    ("Biografilm", "Biografilm Festival"),
    (
        "Mostra Internacional de São Paulo",
        "São Paulo . Mostra Internacional de Cinema",
    ),
    ("IndieLisboa", "IndieLisboa Film Festival"),
    (
        "Alcances",
        "Alcances / Festival de Cine Documental - Sección Oficial (Cádiz - Spain)",
        "Festival de Cine Documental Alcances",
    ),
    ("Festival del film Locarno", "Locarno Film Festival"),
    ("CANNES INTERNATIONAL FILM FESTIVAL", "Festival de Cannes"),
    ("Beldocs", "Beldocs (International Documentary Film Festival)"),
    (
        "Festival Internacional de Cine de Guadalajara (FICG)",
        "Festival Internacional de Cine de Guadalajara",
    ),
    (
        "Festival Internacional de Cine de Cartagena de Indias (FICCI)",
        "Festival Internacional de Cine de Cartagena de Indias",
    ),
    (
        "HOT DOCS CANADIAN INTERNATIONAL DOCUMENTARY FESTIVAL (Canada)",
        "Hot Docs Canadian International Documentary Festival",
    ),
    (
        "BAFICI - Buenos Aires Festival Internacional de Cine Independiente (para 2026)",
        "BAFICI - Buenos Aires Festival Internacional de Cine Independiente",
    ),
    ("Festival de cine de Morelia", "Festival de cine de Morelia (FICM)"),
    (
        "FICINDIE Festival Internacional de Cine Independiente y de Autor de Canarias",
        "Festival Internacional de Cine Independiente y de Autor Canarias",
    ),
    (
        "Los Trabajos y las Noches - Festival de Cine y Procesos Artísticos",
        "Los Trabajos y las Noches - Festival de Cine y Procesos Artísticos "
        "(Logroño - Spain)",
    ),
    ("Another Way Film Festival", "Another Way Film Festival (Spain)"),
    (
        "Festival du Film Hispanique",
        "Festival du Film Hispanique (Le Mans - France)",
    ),
    (
        "International Short Film Festival Oberhausen",
        "Internationale Kurzfilmtage Oberhausen",
    ),
    ("INDIE SHORTS AWARDS CANNES", "Cannes Indie Short Awards"),
    (
        "IBIZACINEFEST-Ibiza Independent Film Festival",
        "Ibiza Cine Fest",
        "Ibiza Cine Fest (Ibiza - Spain)",
    ),
    ("Mirades Fest", "Mirades Fest (para 2026)"),
    ("SHEFFIELD DOC/FEST", "Sheffield International Documentary Festival"),
    (
        "Cinespaña Toulouse",
        "Cinespaña, Festival de Cinéma Espagnol & Portugais de Toulouse",
        "Festival Cinespaña - Documentary Official Competition (Toulouse - France)",
    ),
    (
        "Clermont Ferrand",
        "Festival International du Court Métrage à Clermont-Ferrand",
    ),
    (
        "Premios Pavez",
        "Premios Pávez - Festival Internacional de Talavera de la Reina",
    ),
    (
        "International Festival of Documentary and Short Film of Bilbao – ZINEBI",
        "Zinebi",
    ),
    (
        "Festival de Cine Comprometido Guadalajara",
        "Festival de Cine Solidario de Guadalajara. FESCIGU",
    ),
    (
        "Festival Jóvenes Realizadores",
        "Festival Jóvenes Realizadores Granada",
    ),
    ("SEMINCI", "Semana Internacional de Cine de Valladolid (SEMINCI)"),
    ("Raindance Film Festival", "Radiance Film Festival"),
)


def normalize_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_key(value) -> str:
    text = unicodedata.normalize("NFKD", normalize_text(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def canonical_name(value) -> str:
    text = normalize_key(value)
    text = re.sub(r"\b(?:19|20)\d{2}\b", " ", text)
    text = re.sub(r"\bpara\s+(?:19|20)\d{2}\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


ALIAS_TO_CANONICAL = {
    normalize_key(alias): normalize_key(group[0])
    for group in ALIAS_GROUPS
    for alias in group
}


def semantic_name(value) -> str:
    normalized = normalize_key(value)
    return ALIAS_TO_CANONICAL.get(normalized, normalized)


def invalid_festival_name(value) -> bool:
    normalized = normalize_key(value)
    return (
        normalized in INVALID_NAME_KEYS
        or normalized.startswith("http")
    )


def comparison_tokens(value) -> set[str]:
    return {
        token
        for token in canonical_name(value).split()
        if token not in GENERIC_NAME_WORDS and len(token) > 1
    }


def website_host(value) -> str:
    raw = normalize_text(value).lower()
    if not raw or raw in {"-", "false", "0"}:
        return ""
    if "@" in raw and not raw.startswith(("http://", "https://")):
        return ""
    candidate = raw if "://" in raw else f"https://{raw}"
    try:
        host = urlparse(candidate).netloc.split("@")[-1].split(":")[0]
    except ValueError:
        return ""
    return host.removeprefix("www.")


def date_value(value):
    if value is None or normalize_text(value) in {"", "-", "false", "0"}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = normalize_text(value)
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    return None


def current_importer_accepts_date(value) -> bool:
    if value is None or normalize_text(value) == "":
        return True
    if isinstance(value, (date, datetime, int, float)):
        return True
    return any(
        _can_parse(normalize_text(value), date_format)
        for date_format in DATE_FORMATS
    )


def _can_parse(value: str, date_format: str) -> bool:
    try:
        datetime.strptime(value, date_format)
        return True
    except ValueError:
        return False


def expected_status(opening, deadline, stored_status=None) -> str:
    if normalize_text(stored_status).upper() == "ARCHIVED":
        return "ARCHIVED"
    today = date.today()
    opening_date = date_value(opening)
    deadline_date = date_value(deadline)
    if deadline_date and deadline_date < today:
        return "CLOSED"
    if opening_date and opening_date > today:
        return "UPCOMING"
    if opening_date and deadline_date and opening_date <= today <= deadline_date:
        return "OPEN"
    if deadline_date and deadline_date >= today:
        return "OPEN"
    return "UNKNOWN"


def row_is_empty(row) -> bool:
    return not any(normalize_text(value) for value in row)


def row_is_repeated_header(row) -> bool:
    name = normalize_key(row[0] if row else None)
    return name in HEADER_NAMES


def cell(row, index):
    return row[index] if 0 <= index < len(row) else None


def make_record(sheet, row_number, row, columns):
    return {
        "source": "excel",
        "sheet": sheet,
        "row": row_number,
        "name": normalize_text(cell(row, columns.get("name", 0))),
        "country": normalize_text(cell(row, columns.get("country", -1))),
        "status_raw": normalize_text(cell(row, columns.get("status", -1))),
        "event_date_raw": cell(row, columns.get("event_date", -1)),
        "opening_date_raw": cell(row, columns.get("opening_date", -1)),
        "deadline_raw": cell(row, columns.get("deadline", -1)),
        "website": normalize_text(cell(row, columns.get("website", -1))),
        "contact": normalize_text(cell(row, columns.get("contact", -1))),
    }


def excel_records(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    records = []
    ignored = []
    sheet_stats = {}
    layouts = {
        "Festivals 2025": {
            "name": 0,
            "status": 2,
            "country": 3,
            "event_date": 10,
            "opening_date": 12,
            "deadline": 16,
            "contact": 20,
            "website": 21,
        },
        "Los Días Azules ": {
            "name": 0,
            "status": 2,
            "country": 3,
            "event_date": 7,
            "opening_date": 9,
            "deadline": 13,
            "contact": 17,
            "website": 18,
        },
        "Copia de Los Días Azules ": {
            "name": 0,
            "country": 1,
            "status": 2,
            "event_date": 8,
            "opening_date": 9,
            "deadline": 10,
            "contact": 11,
            "website": 12,
        },
        "AMILCAR Festivals 2025": {
            "name": 0,
            "status": 2,
            "country": 3,
            "event_date": 6,
            "opening_date": 8,
            "deadline": 12,
            "contact": 16,
            "website": 17,
        },
        "subvencionados ICAA": {
            "name": 0,
            "country": 1,
        },
        "Listado festivales": {
            "name": 0,
            "country": 1,
            "status": 2,
            "event_date": 4,
            "opening_date": 5,
            "deadline": 6,
            "contact": 8,
            "website": 9,
        },
    }
    for sheet in workbook.worksheets:
        layout = layouts[sheet.title]
        total_rows = 0
        nonempty_rows = 0
        valid_rows = 0
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            total_rows += 1
            if row_is_empty(row):
                ignored.append((sheet.title, row_number, "empty"))
                continue
            nonempty_rows += 1
            if row_is_repeated_header(row):
                ignored.append((sheet.title, row_number, "header"))
                continue
            name = normalize_text(cell(row, 0))
            country = normalize_text(cell(row, 1))
            if sheet.title == "AMILCAR Festivals 2025" and row_number == 1:
                record = {
                    "source": "excel",
                    "sheet": sheet.title,
                    "row": row_number,
                    "name": name,
                    "country": "",
                    "status_raw": "",
                    "event_date_raw": None,
                    "opening_date_raw": None,
                    "deadline_raw": None,
                    "website": "",
                    "contact": "",
                }
                records.append(record)
                valid_rows += 1
                continue
            if sheet.title == "subvencionados ICAA":
                if not name:
                    ignored.append((sheet.title, row_number, "section_or_note"))
                    continue
            elif not name:
                ignored.append((sheet.title, row_number, "auxiliary"))
                continue
            record = make_record(sheet.title, row_number, row, layout)
            if invalid_festival_name(record["name"]):
                ignored.append((sheet.title, row_number, "invalid_name"))
                continue
            records.append(record)
            valid_rows += 1
        sheet_stats[sheet.title] = {
            "physical_rows": total_rows,
            "nonempty_rows": nonempty_rows,
            "valid_festival_rows": valid_rows,
            "ignored_nonempty_rows": nonempty_rows - valid_rows,
        }
    workbook.close()
    return records, ignored, sheet_stats


def firestore_records():
    from app.core.firebase import db

    records = []
    for snapshot in db.collection("festivals").stream(timeout=60):
        data = snapshot.to_dict() or {}
        records.append(
            {
                "source": "firestore",
                "id": snapshot.id,
                **data,
            }
        )
    return records


def group_duplicates(records, key_function):
    groups = defaultdict(list)
    for record in records:
        key = key_function(record)
        if key:
            groups[key].append(record)
    return {
        key: values
        for key, values in groups.items()
        if len(values) > 1
    }


def identity_key(record):
    name = canonical_name(record.get("name"))
    country = normalize_key(record.get("country"))
    return name, country


def candidate_variant_groups(records):
    by_canonical = group_duplicates(records, lambda record: canonical_name(record["name"]))
    candidates = []
    seen = set()
    for values in by_canonical.values():
        names = sorted({record["name"] for record in values})
        if len(names) > 1:
            candidates.append(values)
            seen.update(canonical_name(record["name"]) for record in values)

    unique_names = {}
    for record in records:
        unique_names.setdefault(canonical_name(record["name"]), record)
    keys = sorted(unique_names)
    for index, left_key in enumerate(keys):
        left_tokens = comparison_tokens(left_key)
        if not left_tokens:
            continue
        for right_key in keys[index + 1:]:
            if left_key == right_key:
                continue
            right_tokens = comparison_tokens(right_key)
            if not right_tokens:
                continue
            overlap = len(left_tokens & right_tokens) / len(left_tokens | right_tokens)
            ratio = SequenceMatcher(None, left_key, right_key).ratio()
            left = unique_names[left_key]
            right = unique_names[right_key]
            same_host = (
                website_host(left.get("website"))
                and website_host(left.get("website")) == website_host(right.get("website"))
            )
            if same_host or overlap >= 0.8 or ratio >= 0.9:
                pair_key = tuple(sorted((left_key, right_key)))
                if pair_key not in seen:
                    candidates.append([left, right])
                    seen.add(pair_key)
    return candidates


def duplicate_summary(groups):
    return {
        "groups": len(groups),
        "rows_in_groups": sum(len(values) for values in groups.values()),
        "duplicate_rows": sum(len(values) - 1 for values in groups.values()),
    }


def display_record(record):
    return {
        key: record.get(key, "")
        for key in ("id", "sheet", "row", "name", "country", "edition_year")
        if record.get(key, "") != ""
    }


def incomplete_fields(record):
    required = ("name", "country", "website", "opening_date", "deadline", "event_date")
    return [
        field
        for field in required
        if not normalize_text(record.get(field))
    ]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", type=Path)
    parser.add_argument("--firestore", action="store_true")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    excel, ignored, sheets = excel_records(args.excel)
    exact = group_duplicates(excel, lambda record: normalize_text(record["name"]))
    normalized = group_duplicates(excel, lambda record: normalize_key(record["name"]))
    canonical = group_duplicates(excel, lambda record: canonical_name(record["name"]))
    semantic = group_duplicates(excel, lambda record: semantic_name(record["name"]))
    identity = group_duplicates(excel, identity_key)
    date_errors = []
    for record in excel:
        for field in ("event_date_raw", "opening_date_raw", "deadline_raw"):
            value = record[field]
            if not current_importer_accepts_date(value):
                date_errors.append(
                    {
                        **display_record(record),
                        "field": field.removesuffix("_raw"),
                        "value": normalize_text(value),
                    }
                )

    report = {
        "excel": {
            "file": str(args.excel),
            "sheets": sheets,
            "total_physical_rows": sum(item["physical_rows"] for item in sheets.values()),
            "total_nonempty_rows": sum(item["nonempty_rows"] for item in sheets.values()),
            "total_valid_rows": len(excel),
            "ignored_nonempty_rows": len(
                [item for item in ignored if item[2] != "empty"]
            ),
            "duplicates_exact_name": duplicate_summary(exact),
            "duplicates_normalized_name": duplicate_summary(normalized),
            "duplicates_canonical_name": duplicate_summary(canonical),
            "duplicates_identity_name_country": duplicate_summary(identity),
            "duplicates_semantic_name": duplicate_summary(semantic),
            "unique_normalized_names": len(
                {normalize_key(record["name"]) for record in excel}
            ),
            "unique_canonical_names": len(
                {canonical_name(record["name"]) for record in excel}
            ),
            "unique_identity_name_country": len(
                {identity_key(record) for record in excel}
            ),
            "unique_semantic_names": len(
                {semantic_name(record["name"]) for record in excel}
            ),
            "date_errors_current_importer": date_errors,
            "exact_duplicate_examples": [
                [display_record(record) for record in values]
                for values in sorted(exact.values(), key=len, reverse=True)[:30]
            ],
        }
    }

    if args.firestore:
        all_firestore = firestore_records()
        invalid_firestore = [
            record
            for record in all_firestore
            if invalid_festival_name(record.get("name"))
        ]
        firestore = [
            record
            for record in all_firestore
            if not invalid_festival_name(record.get("name"))
        ]
        fs_exact = group_duplicates(
            firestore, lambda record: normalize_text(record.get("name"))
        )
        fs_normalized = group_duplicates(
            firestore, lambda record: normalize_key(record.get("name"))
        )
        fs_identity = group_duplicates(firestore, identity_key)
        fs_semantic = group_duplicates(
            firestore, lambda record: semantic_name(record.get("name"))
        )
        excel_keys = {semantic_name(record["name"]) for record in excel}
        firestore_keys = {
            semantic_name(record.get("name"))
            for record in firestore
        }
        status_raw = Counter(
            normalize_text(record.get("status")).upper() or "UNKNOWN"
            for record in firestore
        )
        status_raw_all = Counter(
            normalize_text(record.get("status")).upper() or "UNKNOWN"
            for record in all_firestore
        )
        status_normalized = Counter(
            status if status in STATUS_VALUES else "UNKNOWN"
            for status in (
                normalize_text(record.get("status")).upper()
                for record in firestore
            )
        )
        status_expected = Counter(
            expected_status(
                record.get("opening_date"),
                record.get("deadline"),
                record.get("status"),
            )
            for record in firestore
        )
        incomplete = [
            {
                **display_record(record),
                "missing": incomplete_fields(record),
            }
            for record in firestore
            if incomplete_fields(record)
        ]
        report["firestore"] = {
            "total_documents_raw": len(all_firestore),
            "total_documents_valid": len(firestore),
            "invalid_documents": [
                display_record(record)
                for record in invalid_firestore
            ],
            "unique_normalized_names": len(
                {normalize_key(record.get("name")) for record in firestore}
            ),
            "unique_identity_name_country": len(firestore_keys),
            "duplicates_exact_name": duplicate_summary(fs_exact),
            "duplicates_normalized_name": duplicate_summary(fs_normalized),
            "duplicates_identity_name_country": duplicate_summary(fs_identity),
            "duplicates_semantic_name": duplicate_summary(fs_semantic),
            "unique_semantic_names": len(firestore_keys),
            "status_raw": dict(sorted(status_raw.items())),
            "status_raw_all_documents": dict(sorted(status_raw_all.items())),
            "status_normalized": dict(sorted(status_normalized.items())),
            "status_recalculated": dict(sorted(status_expected.items())),
            "incomplete_records": len(incomplete),
            "incomplete_by_field": dict(
                Counter(field for item in incomplete for field in item["missing"])
            ),
            "incomplete_examples": incomplete[:50],
            "duplicate_examples": [
                [display_record(record) for record in values]
                for values in sorted(fs_identity.values(), key=len, reverse=True)[:50]
            ],
            "semantic_duplicate_groups": [
                [display_record(record) for record in values]
                for values in sorted(
                    fs_semantic.values(),
                    key=lambda values: (-len(values), semantic_name(values[0]["name"])),
                )
            ],
            "missing_from_firestore": [
                display_record(record)
                for record in excel
                if semantic_name(record["name"]) not in firestore_keys
            ],
            "extra_in_firestore": [
                display_record(record)
                for record in firestore
                if semantic_name(record.get("name")) not in excel_keys
            ],
            "missing_unique_count": len(excel_keys - firestore_keys),
            "extra_unique_count": len(firestore_keys - excel_keys),
        }

    output = json.dumps(report, ensure_ascii=False, indent=2, default=str)
    if args.output:
        args.output.write_text(output, encoding="utf-8")
    else:
        print(output)


if __name__ == "__main__":
    main()
